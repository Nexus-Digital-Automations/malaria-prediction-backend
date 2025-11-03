"""
Report Generation Engine for Custom Reports and Export System.

This module provides comprehensive report generation capabilities with support
for multiple export formats, template processing, chart rendering, and automated
scheduling. Integrates with existing analytics infrastructure for data access.
"""

import base64
import json
import logging
import os
import tempfile
import time
from datetime import datetime
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any

import matplotlib.pyplot as plt
import pandas as pd
import plotly.graph_objects as go
from jinja2 import Environment, FileSystemLoader, Template
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (
    Image as RLImage,
)
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
)
from sqlalchemy.orm import Session

from ..database.models import (
    ERA5DataPoint,
    MalariaRiskIndex,
    Report,
    ReportTemplate,
)
from .data_export import DataExportService

logger = logging.getLogger(__name__)


class ChartRenderer:
    """High-quality chart rendering for multiple export formats."""

    def __init__(self) -> None:
        """Initialize chart renderer with default settings."""
        self.dpi = 300  # High DPI for quality exports
        self.default_style = {
            'figure.figsize': (10, 6),
            'font.size': 10,
            'axes.titlesize': 12,
            'axes.labelsize': 10,
            'xtick.labelsize': 9,
            'ytick.labelsize': 9,
            'legend.fontsize': 9,
        }

    def create_time_series_chart(  # type: ignore[no-untyped-def]
        self,
        data: pd.DataFrame,
        x_column: str,
        y_columns: list[str],
        title: str,
        chart_type: str = "line",
        **kwargs
    ) -> dict[str, Any]:
        """
        Create time series chart for reports.

        Args:
            data: DataFrame with time series data
            x_column: Column name for x-axis (typically date/time)
            y_columns: List of column names for y-axis
            title: Chart title
            chart_type: Type of chart (line, bar, area)
            **kwargs: Additional chart configuration

        Returns:
            Dictionary with chart data for multiple formats
        """
        try:
            plt.style.use('seaborn-v0_8')
            plt.rcParams.update(self.default_style)

            fig, ax = plt.subplots(figsize=kwargs.get('figsize', (10, 6)))

            # Plot based on chart type
            if chart_type == "line":
                for column in y_columns:
                    ax.plot(data[x_column], data[column], label=column, linewidth=2)
            elif chart_type == "bar":
                data.set_index(x_column)[y_columns].plot(kind='bar', ax=ax, width=0.8)
            elif chart_type == "area":
                data.set_index(x_column)[y_columns].plot(kind='area', ax=ax, alpha=0.7)

            # Styling
            ax.set_title(title, fontsize=14, fontweight='bold', pad=20)
            ax.set_xlabel(kwargs.get('xlabel', x_column.replace('_', ' ').title()), fontsize=11)
            ax.set_ylabel(kwargs.get('ylabel', 'Value'), fontsize=11)
            ax.legend(bbox_to_anchor=(1.05, 1), loc='upper left')
            ax.grid(True, alpha=0.3)

            # Format dates if x-axis is datetime
            if pd.api.types.is_datetime64_any_dtype(data[x_column]):
                fig.autofmt_xdate()

            plt.tight_layout()

            # Generate multiple format outputs
            chart_data = {}

            # PNG for general use
            png_buffer = BytesIO()
            plt.savefig(png_buffer, format='png', dpi=self.dpi, bbox_inches='tight')
            png_buffer.seek(0)
            chart_data['png'] = base64.b64encode(png_buffer.getvalue()).decode()

            # SVG for vector graphics
            svg_buffer = BytesIO()
            plt.savefig(svg_buffer, format='svg', bbox_inches='tight')
            svg_buffer.seek(0)
            chart_data['svg'] = svg_buffer.getvalue().decode()

            # Plotly for interactive charts
            plotly_fig = self._create_plotly_chart(data, x_column, y_columns, title, chart_type, **kwargs)
            chart_data['plotly_json'] = plotly_fig.to_json()
            chart_data['plotly_html'] = plotly_fig.to_html(include_plotlyjs=True)

            plt.close(fig)
            return chart_data

        except Exception as e:
            logger.error(f"Error creating time series chart: {str(e)}")
            return {}

    def _create_plotly_chart(  # type: ignore[no-untyped-def]
        self,
        data: pd.DataFrame,
        x_column: str,
        y_columns: list[str],
        title: str,
        chart_type: str,
        **kwargs
    ) -> go.Figure:
        """Create interactive Plotly chart."""
        fig = go.Figure()

        for column in y_columns:
            if chart_type == "line":
                fig.add_trace(go.Scatter(
                    x=data[x_column],
                    y=data[column],
                    mode='lines+markers',
                    name=column,
                    line={'width': 2}
                ))
            elif chart_type == "bar":
                fig.add_trace(go.Bar(
                    x=data[x_column],
                    y=data[column],
                    name=column
                ))
            elif chart_type == "area":
                fig.add_trace(go.Scatter(
                    x=data[x_column],
                    y=data[column],
                    fill='tonexty' if column != y_columns[0] else 'tozeroy',
                    name=column
                ))

        fig.update_layout(
            title={'text': title, 'font': {'size': 16, 'family': "Arial, sans-serif"}},
            xaxis_title=kwargs.get('xlabel', x_column.replace('_', ' ').title()),
            yaxis_title=kwargs.get('ylabel', 'Value'),
            hovermode='x unified',
            showlegend=True,
            plot_bgcolor='white',
            paper_bgcolor='white',
            margin={'l': 50, 'r': 50, 't': 50, 'b': 50}
        )

        fig.update_xaxes(showgrid=True, gridwidth=1, gridcolor='lightgray')
        fig.update_yaxes(showgrid=True, gridwidth=1, gridcolor='lightgray')

        return fig

    def create_distribution_chart(  # type: ignore[no-untyped-def]
        self,
        data: pd.DataFrame,
        column: str,
        title: str,
        chart_type: str = "histogram",
        **kwargs
    ) -> dict[str, Any]:
        """Create distribution chart (histogram, box plot, etc.)."""
        try:
            plt.style.use('seaborn-v0_8')
            plt.rcParams.update(self.default_style)

            fig, ax = plt.subplots(figsize=kwargs.get('figsize', (8, 6)))

            if chart_type == "histogram":
                ax.hist(data[column], bins=kwargs.get('bins', 30), alpha=0.7, edgecolor='black')
                ax.set_ylabel('Frequency')
            elif chart_type == "box":
                ax.boxplot(data[column])
                ax.set_ylabel('Value')

            ax.set_title(title, fontsize=14, fontweight='bold', pad=20)
            ax.set_xlabel(column.replace('_', ' ').title())
            ax.grid(True, alpha=0.3)

            plt.tight_layout()

            # Generate outputs
            chart_data = {}
            png_buffer = BytesIO()
            plt.savefig(png_buffer, format='png', dpi=self.dpi, bbox_inches='tight')
            png_buffer.seek(0)
            chart_data['png'] = base64.b64encode(png_buffer.getvalue()).decode()

            plt.close(fig)
            return chart_data

        except Exception as e:
            logger.error(f"Error creating distribution chart: {str(e)}")
            return {}


class TemplateProcessor:
    """Template processing engine for report generation."""

    def __init__(self, template_dir: str | None = None) -> None:
        """
        Initialize template processor.

        Args:
            template_dir: Directory containing report templates
        """
        self.template_dir = template_dir or self._get_default_template_dir()
        self.env = Environment(
            loader=FileSystemLoader(self.template_dir),
            autoescape=True
        )

    def _get_default_template_dir(self) -> str:
        """Get default template directory."""
        current_dir = Path(__file__).parent
        template_dir = current_dir / "templates" / "reports"
        template_dir.mkdir(parents=True, exist_ok=True)
        return str(template_dir)

    def process_template(
        self,
        template_config: dict[str, Any],
        data: dict[str, Any],
        charts: dict[str, Any]
    ) -> dict[str, Any]:
        """
        Process report template with data and charts.

        Args:
            template_config: Template configuration from database
            data: Report data to insert
            charts: Chart data for inclusion

        Returns:
            Processed template content for multiple formats
        """
        try:
            # Extract template components
            layout = template_config.get('layout_configuration', {})
            widgets = template_config.get('widgets', [])
            style = template_config.get('style_configuration', {})

            # Process each widget
            processed_widgets = []
            for widget in widgets:
                processed_widget = self._process_widget(widget, data, charts)
                processed_widgets.append(processed_widget)

            # Generate template outputs
            return {
                'html': self._generate_html_template(layout, processed_widgets, style),
                'markdown': self._generate_markdown_template(processed_widgets),
                'json': self._generate_json_template(layout, processed_widgets),
                'widgets': processed_widgets
            }

        except Exception as e:
            logger.error(f"Error processing template: {str(e)}")
            return {}

    def _process_widget(
        self,
        widget: dict[str, Any],
        data: dict[str, Any],
        charts: dict[str, Any]
    ) -> dict[str, Any]:
        """Process individual widget with data."""
        widget_type = widget.get('type', 'text')
        widget_id = widget.get('id', '')

        if widget_type == 'chart':
            chart_id = widget.get('chart_id', '')
            return {
                'type': 'chart',
                'id': widget_id,
                'title': widget.get('title', ''),
                'chart_data': charts.get(chart_id, {}),
                'config': widget.get('config', {})
            }
        elif widget_type == 'table':
            data_source = widget.get('data_source', '')
            return {
                'type': 'table',
                'id': widget_id,
                'title': widget.get('title', ''),
                'data': data.get(data_source, []),
                'config': widget.get('config', {})
            }
        elif widget_type == 'text':
            return {
                'type': 'text',
                'id': widget_id,
                'title': widget.get('title', ''),
                'content': widget.get('content', ''),
                'config': widget.get('config', {})
            }
        else:
            return widget

    def _generate_html_template(
        self,
        layout: dict[str, Any],
        widgets: list[dict[str, Any]],
        style: dict[str, Any]
    ) -> str:
        """Generate HTML template."""
        html_template = """
        <!DOCTYPE html>
        <html>
        <head>
            <title>{{ title }}</title>
            <style>
                body { font-family: Arial, sans-serif; margin: 20px; }
                .widget { margin: 20px 0; padding: 15px; border: 1px solid #ddd; }
                .chart { text-align: center; }
                .table { width: 100%; border-collapse: collapse; }
                .table th, .table td { border: 1px solid #ddd; padding: 8px; text-align: left; }
                .table th { background-color: #f2f2f2; }
            </style>
        </head>
        <body>
            <h1>{{ title }}</h1>
            {% for widget in widgets %}
                <div class="widget">
                    {% if widget.title %}<h2>{{ widget.title }}</h2>{% endif %}
                    {% if widget.type == 'chart' %}
                        <div class="chart">{{ widget.chart_data.plotly_html | safe }}</div>
                    {% elif widget.type == 'table' %}
                        <table class="table">
                            {% if widget.data %}
                                {% for row in widget.data[:1] %}
                                <tr>{% for key in row.keys() %}<th>{{ key }}</th>{% endfor %}</tr>
                                {% endfor %}
                                {% for row in widget.data %}
                                <tr>{% for value in row.values() %}<td>{{ value }}</td>{% endfor %}</tr>
                                {% endfor %}
                            {% endif %}
                        </table>
                    {% elif widget.type == 'text' %}
                        <p>{{ widget.content }}</p>
                    {% endif %}
                </div>
            {% endfor %}
        </body>
        </html>
        """
        template = Template(html_template)
        return template.render(
            title=layout.get('title', 'Report'),
            widgets=widgets
        )

    def _generate_markdown_template(self, widgets: list[dict[str, Any]]) -> str:
        """Generate Markdown template."""
        markdown_lines = ["# Report\n"]

        for widget in widgets:
            if widget.get('title'):
                markdown_lines.append(f"## {widget['title']}\n")

            if widget['type'] == 'text':
                markdown_lines.append(f"{widget.get('content', '')}\n")
            elif widget['type'] == 'table':
                # Simple table formatting
                data = widget.get('data', [])
                if data:
                    headers = list(data[0].keys())
                    markdown_lines.append("| " + " | ".join(headers) + " |")
                    markdown_lines.append("|" + "---|" * len(headers))
                    for row in data[:10]:  # Limit to 10 rows
                        values = [str(row.get(h, '')) for h in headers]
                        markdown_lines.append("| " + " | ".join(values) + " |")
                markdown_lines.append("")

        return "\n".join(markdown_lines)

    def _generate_json_template(
        self,
        layout: dict[str, Any],
        widgets: list[dict[str, Any]]
    ) -> str:
        """Generate JSON template."""
        return json.dumps({
            'layout': layout,
            'widgets': widgets,
            'generated_at': datetime.utcnow().isoformat()
        }, indent=2)


class ReportExporter:
    """Multi-format report export engine."""

    def __init__(self, temp_dir: str | None = None) -> None:
        """
        Initialize report exporter.

        Args:
            temp_dir: Temporary directory for file generation
        """
        self.temp_dir = temp_dir or tempfile.gettempdir()
        self.chart_renderer = ChartRenderer()

    async def export_to_pdf(
        self,
        template_content: str,
        charts: dict[str, Any],
        metadata: dict[str, Any]
    ) -> bytes:
        """
        Export report to PDF format.

        Args:
            template_content: Processed template content
            charts: Chart data
            metadata: Report metadata

        Returns:
            PDF file as bytes
        """
        try:
            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=A4)
            story = []
            styles = getSampleStyleSheet()

            # Title
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Title'],
                fontSize=18,
                spaceAfter=30,
                alignment=1  # Center alignment
            )
            story.append(Paragraph(metadata.get('title', 'Report'), title_style))

            # Metadata
            meta_info = f"Generated: {metadata.get('generated_at', datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC'))}"
            story.append(Paragraph(meta_info, styles['Normal']))
            story.append(Spacer(1, 20))

            # Process template content (simplified HTML to PDF conversion)
            content_paragraphs = template_content.split('\n')
            for para in content_paragraphs:
                if para.strip():
                    story.append(Paragraph(para.strip(), styles['Normal']))

            # Add charts
            for chart_id, chart_data in charts.items():
                if 'png' in chart_data:
                    try:
                        # Decode base64 image
                        image_data = base64.b64decode(chart_data['png'])

                        # Create temporary file for image
                        with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as temp_file:
                            temp_file.write(image_data)
                            temp_file.flush()

                            # Add image to PDF
                            img = RLImage(temp_file.name, width=6*inch, height=4*inch)
                            story.append(img)
                            story.append(Spacer(1, 20))

                        # Clean up temporary file
                        os.unlink(temp_file.name)

                    except Exception as e:
                        logger.error(f"Error adding chart {chart_id} to PDF: {str(e)}")

            doc.build(story)
            buffer.seek(0)
            return buffer.getvalue()

        except Exception as e:
            logger.error(f"Error exporting to PDF: {str(e)}")
            raise

    async def export_to_excel(
        self,
        data: dict[str, Any],
        charts: dict[str, Any],
        metadata: dict[str, Any]
    ) -> bytes:
        """
        Export report to Excel format.

        Args:
            data: Report data
            charts: Chart data
            metadata: Report metadata

        Returns:
            Excel file as bytes
        """
        try:
            buffer = BytesIO()
            workbook = Workbook()

            # Remove default sheet
            workbook.remove(workbook.active)

            # Summary sheet
            summary_sheet = workbook.create_sheet("Summary")
            summary_sheet['A1'] = metadata.get('title', 'Report')
            summary_sheet['A1'].font = Font(size=16, bold=True)
            summary_sheet['A3'] = f"Generated: {metadata.get('generated_at', datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC'))}"

            # Data sheets
            for data_name, dataset in data.items():
                if isinstance(dataset, pd.DataFrame):
                    sheet = workbook.create_sheet(data_name[:31])  # Excel sheet name limit

                    # Add data
                    for r in dataframe_to_rows(dataset, index=False, header=True):
                        sheet.append(r)

                    # Format header
                    for cell in sheet[1]:
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                        cell.alignment = Alignment(horizontal="center")

            # Charts sheet
            if charts:
                chart_sheet = workbook.create_sheet("Charts")
                row = 1

                for chart_id, chart_data in charts.items():
                    chart_sheet[f'A{row}'] = f"Chart: {chart_id}"
                    chart_sheet[f'A{row}'].font = Font(bold=True)

                    # Add chart as image if PNG data available
                    if 'png' in chart_data:
                        try:
                            image_data = base64.b64decode(chart_data['png'])
                            with tempfile.NamedTemporaryFile(suffix='.png', delete=False) as temp_file:
                                temp_file.write(image_data)
                                temp_file.flush()

                                img = Image(temp_file.name)
                                img.width = 600
                                img.height = 400
                                chart_sheet.add_image(img, f'A{row + 1}')

                                os.unlink(temp_file.name)

                        except Exception as e:
                            logger.error(f"Error adding chart {chart_id} to Excel: {str(e)}")

                    row += 25  # Space for next chart

            workbook.save(buffer)
            buffer.seek(0)
            return buffer.getvalue()

        except Exception as e:
            logger.error(f"Error exporting to Excel: {str(e)}")
            raise

    async def export_to_csv(
        self,
        data: dict[str, Any],
        metadata: dict[str, Any]
    ) -> bytes:
        """
        Export report data to CSV format.

        Args:
            data: Report data
            metadata: Report metadata

        Returns:
            CSV file as bytes
        """
        try:
            buffer = StringIO()

            # Write metadata header
            buffer.write(f"# {metadata.get('title', 'Report')}\n")
            buffer.write(f"# Generated: {metadata.get('generated_at', datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC'))}\n")
            buffer.write("\n")

            # Write data sections
            for data_name, dataset in data.items():
                buffer.write(f"# {data_name}\n")

                if isinstance(dataset, pd.DataFrame):
                    dataset.to_csv(buffer, index=False)
                elif isinstance(dataset, list):
                    if dataset and isinstance(dataset[0], dict):
                        df = pd.DataFrame(dataset)
                        df.to_csv(buffer, index=False)

                buffer.write("\n")

            content = buffer.getvalue()
            return content.encode('utf-8')

        except Exception as e:
            logger.error(f"Error exporting to CSV: {str(e)}")
            raise

    async def export_to_powerpoint(
        self,
        template_content: str,
        charts: dict[str, Any],
        metadata: dict[str, Any]
    ) -> bytes:
        """
        Export report to PowerPoint format.

        Args:
            template_content: Processed template content
            charts: Chart data
            metadata: Report metadata

        Returns:
            PowerPoint file as bytes
        """
        try:
            # Note: This would require python-pptx library
            # For now, return a placeholder
            logger.warning("PowerPoint export not fully implemented - requires python-pptx")
            return b"PowerPoint export placeholder"

        except Exception as e:
            logger.error(f"Error exporting to PowerPoint: {str(e)}")
            raise


class ReportGenerator:
    """Main report generation engine with comprehensive functionality."""

    def __init__(self, db_session: Session) -> None:
        """
        Initialize report generator.

        Args:
            db_session: Database session for data access
        """
        self.db = db_session
        self.data_export = DataExportService(db_session)  # type: ignore[arg-type]
        self.chart_renderer = ChartRenderer()
        self.template_processor = TemplateProcessor()
        self.exporter = ReportExporter()

    async def generate_report(
        self,
        template_id: int | None = None,
        report_config: dict[str, Any] | None = None,
        user_id: str = "system",
        export_formats: list[str] | None = None
    ) -> dict[str, Any]:
        """
        Generate comprehensive report with multiple export formats.

        Args:
            template_id: ID of report template to use
            report_config: Custom report configuration
            user_id: User generating the report
            export_formats: List of formats to export ["pdf", "excel", "csv", "pptx"]

        Returns:
            Report generation results with file paths and metadata
        """
        start_time = time.time()
        export_formats = export_formats or ["pdf", "excel", "csv"]

        try:
            logger.info(f"Starting report generation for user {user_id}")

            # Load template if specified
            template = None
            if template_id:
                template = self.db.query(ReportTemplate).filter(ReportTemplate.id == template_id).first()
                if not template:
                    raise ValueError(f"Template {template_id} not found")

            # Generate report data
            report_data = await self._generate_report_data(report_config or {})

            # Generate charts
            charts = await self._generate_charts(report_data, template)

            # Process template
            template_config = template.layout_configuration if template else self._get_default_template()
            template_content = self.template_processor.process_template(
                dict(template_config) if not isinstance(template_config, dict) else template_config,
                report_data,
                charts
            )

            # Create report record
            report = Report(
                title=report_config.get('title', 'Generated Report'),  # type: ignore[union-attr]
                description=report_config.get('description'),  # type: ignore[union-attr]
                report_type=report_config.get('type', 'custom'),  # type: ignore[union-attr]
                template_id=template_id,
                generated_by=user_id,
                report_data=report_data,
                chart_configurations=charts,
                custom_parameters=report_config,
                export_formats=export_formats,
                status="generating"
            )
            self.db.add(report)
            self.db.commit()

            # Generate exports
            export_results = {}
            file_paths = {}
            file_sizes = {}

            metadata = {
                'title': report.title,
                'generated_at': report.generated_at.isoformat(),
                'user_id': user_id,
                'report_id': report.id
            }

            for format_name in export_formats:
                try:
                    logger.info(f"Generating {format_name} export")

                    if format_name == "pdf":
                        content = await self.exporter.export_to_pdf(
                            template_content['html'], charts, metadata
                        )
                    elif format_name == "excel":
                        content = await self.exporter.export_to_excel(
                            report_data, charts, metadata
                        )
                    elif format_name == "csv":
                        content = await self.exporter.export_to_csv(
                            report_data, metadata
                        )
                    elif format_name == "pptx":
                        content = await self.exporter.export_to_powerpoint(
                            template_content['html'], charts, metadata
                        )
                    else:
                        logger.warning(f"Unsupported export format: {format_name}")
                        continue

                    # Save file
                    file_path = await self._save_export_file(int(report.id), format_name, content)
                    file_paths[format_name] = file_path
                    file_sizes[format_name] = len(content)
                    export_results[format_name] = {
                        'status': 'completed',
                        'path': file_path,
                        'generated_at': datetime.utcnow().isoformat(),
                        'size_bytes': len(content)
                    }

                    logger.info(f"Generated {format_name} export: {len(content)} bytes")

                except Exception as e:
                    logger.error(f"Error generating {format_name} export: {str(e)}")
                    export_results[format_name] = {
                        'status': 'failed',
                        'error': str(e),
                        'generated_at': datetime.utcnow().isoformat()
                    }

            # Update report with results
            generation_time = time.time() - start_time
            report.status = "completed"  # type: ignore[assignment]
            report.export_status = export_results  # type: ignore[assignment]
            report.file_paths = file_paths  # type: ignore[assignment]
            report.file_sizes = file_sizes  # type: ignore[assignment]
            report.generation_time_seconds = generation_time  # type: ignore[assignment]
            report.data_points_count = self._count_data_points(report_data)  # type: ignore[assignment]

            self.db.commit()

            logger.info(f"Report generation completed in {generation_time:.2f} seconds")

            return {
                'report_id': report.id,
                'status': 'completed',
                'generation_time_seconds': generation_time,
                'export_results': export_results,
                'file_paths': file_paths,
                'data_points_count': report.data_points_count
            }

        except Exception as e:
            logger.error(f"Error generating report: {str(e)}")

            # Update report status if it exists
            if 'report' in locals():
                report.status = "failed"  # type: ignore[assignment]
                report.error_message = str(e)  # type: ignore[assignment]
                self.db.commit()

            raise

    async def _generate_report_data(self, config: dict[str, Any]) -> dict[str, Any]:
        """Generate report data based on configuration."""
        data = {}

        # Prediction accuracy data
        if config.get('include_prediction_accuracy', True):
            start_date = config.get('start_date')
            end_date = config.get('end_date')

            query = self.db.query(MalariaRiskIndex)
            if start_date:
                query = query.filter(MalariaRiskIndex.assessment_date >= start_date)
            if end_date:
                query = query.filter(MalariaRiskIndex.assessment_date <= end_date)
            predictions = query.limit(1000).all()

            data['predictions'] = [
                {
                    'date': p.assessment_date.isoformat(),
                    'location': p.location_name,
                    'risk_level': p.risk_level,
                    'confidence': p.confidence_score,
                    'model_type': p.model_type
                }
                for p in predictions
            ]

        # Climate data
        if config.get('include_climate_data', True):
            climate_data = self.db.query(ERA5DataPoint).limit(500).all()
            data['climate'] = [
                {
                    'date': c.timestamp.isoformat(),
                    'latitude': c.latitude,
                    'longitude': c.longitude,
                    'temperature': c.temperature_2m,
                    'precipitation': c.total_precipitation
                }
                for c in climate_data
            ]

        return data

    async def _generate_charts(
        self,
        data: dict[str, Any],
        template: ReportTemplate | None
    ) -> dict[str, Any]:
        """Generate charts based on data and template configuration."""
        charts = {}

        # Prediction trends chart
        if 'predictions' in data:
            df = pd.DataFrame(data['predictions'])
            if not df.empty:
                df['date'] = pd.to_datetime(df['date'])

                # Risk level distribution
                charts['risk_distribution'] = self.chart_renderer.create_distribution_chart(
                    df, 'risk_level', 'Risk Level Distribution'
                )

                # Confidence over time
                if len(df) > 1:
                    charts['confidence_trend'] = self.chart_renderer.create_time_series_chart(
                        df, 'date', ['confidence'], 'Prediction Confidence Over Time'
                    )

        # Climate trends
        if 'climate' in data:
            df = pd.DataFrame(data['climate'])
            if not df.empty:
                df['date'] = pd.to_datetime(df['date'])

                charts['temperature_trend'] = self.chart_renderer.create_time_series_chart(
                    df, 'date', ['temperature'], 'Temperature Trends'
                )

                charts['precipitation_trend'] = self.chart_renderer.create_time_series_chart(
                    df, 'date', ['precipitation'], 'Precipitation Trends'
                )

        return charts

    def _get_default_template(self) -> dict[str, Any]:
        """Get default template configuration."""
        return {
            'layout_configuration': {
                'title': 'Malaria Prediction Report',
                'orientation': 'portrait',
                'sections': ['summary', 'charts', 'data']
            },
            'widgets': [
                {'type': 'text', 'id': 'summary', 'title': 'Summary', 'content': 'Report summary'},
                {'type': 'chart', 'id': 'risk_chart', 'title': 'Risk Analysis', 'chart_id': 'risk_distribution'},
                {'type': 'table', 'id': 'data_table', 'title': 'Data', 'data_source': 'predictions'}
            ]
        }

    async def _save_export_file(self, report_id: int, format_name: str, content: bytes) -> str:
        """Save exported file and return relative path."""
        # Create reports directory
        reports_dir = Path("exports/reports")
        reports_dir.mkdir(parents=True, exist_ok=True)

        # Generate filename
        timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        filename = f"report_{report_id}_{timestamp}.{format_name}"
        file_path = reports_dir / filename

        # Save file
        with open(file_path, 'wb') as f:
            f.write(content)

        return str(file_path)

    def _count_data_points(self, data: dict[str, Any]) -> int:
        """Count total data points in report."""
        total = 0
        for dataset in data.values():
            if isinstance(dataset, list):
                total += len(dataset)
            elif isinstance(dataset, pd.DataFrame):
                total += len(dataset)
        return total


# Service factory function
def get_report_generator(db_session: Session) -> ReportGenerator:
    """
    Factory function to create ReportGenerator instance.

    Args:
        db_session: Database session

    Returns:
        Configured ReportGenerator instance
    """
    return ReportGenerator(db_session)
